# import numpy as np
# import matplotlib as plt
import openpyxl as op

DataAll = op.load_workbook("results.xlsx")
data = DataAll['Sheet1']

r = 1
Reward = 1
alpha = 0.2

result_my = [[0 for i in range(20)] for j in range(20)]
result_oppo = [[0 for i in range(20)] for j in range(20)]

FVMC = 1  # First-Visit Monto Carlo (MC) on policy evaluation
IMC = 0  # Temporal Difference Policy Evaluation
################################################################
if (FVMC and IMC) or ((not FVMC) and (not IMC)):
    raise Exception

################################################################
if FVMC:
    for TEST_NUM in range(1 + 1, 132 + 1):
        reward = 1
        STEP_NUM = 4
        PLAY_FIRST = 2 - data.cell(TEST_NUM, 2).value
        while True:
            if data.cell(TEST_NUM, STEP_NUM).value is None:
                break
            else:
                x, y = eval(data.cell(TEST_NUM, STEP_NUM).value)
                if STEP_NUM % 2 == 1 - PLAY_FIRST:
                    result_my[x][y] += pow(r, STEP_NUM)
                else:
                    result_oppo[x][y] += pow(r, STEP_NUM)
            STEP_NUM += 1

#################################################################
if IMC:
    for TEST_NUM in range(1 + 1, 132 + 1):
        reward = 1
        STEP_NUM = 4
        PLAY_FIRST = 2 - data.cell(TEST_NUM, 2).value
        while True:
            if data.cell(TEST_NUM, STEP_NUM).value is None:
                break
            else:
                x, y = eval(data.cell(TEST_NUM, STEP_NUM).value)
                if STEP_NUM % 2 == 1 - PLAY_FIRST:
                    result_my[x][y] = (1 - alpha) * result_my[x][y] + alpha * pow(r, STEP_NUM)
                else:
                    result_oppo[x][y] = (1 - alpha) * result_my[x][y] + alpha * pow(r, STEP_NUM)
            STEP_NUM += 1

###########################################################################
